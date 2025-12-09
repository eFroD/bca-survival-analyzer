#!/usr/bin/env python3
import argparse
import os
import sys
from typing import Any, Dict, List, Literal, Optional, Type, cast

import pandas as pd

# Define PDF converter type
PDF_CONVERTER_TYPE = Literal["win32", "fpdf", None]
PDF_CONVERTER: PDF_CONVERTER_TYPE = None

# Define base FPDF type for later conditional definition
FPDF_CLASS: Optional[Type[Any]] = None
PDFWithWordWrap_CLASS: Optional[Type[Any]] = None

# For PDF conversion
try:
    # Windows approach using COM automation
    import win32com.client  # type: ignore[import-untyped]

    PDF_CONVERTER = "win32"
except ImportError:
    try:
        # Alternative approach using fpdf
        import openpyxl  # type: ignore[import-untyped]
        from fpdf import FPDF  # type: ignore[import-untyped]

        # Store the FPDF class for later use
        FPDF_CLASS = FPDF

        class PDFWithWordWrap(FPDF):  # type: ignore[valid-type]
            """Extended FPDF class with word wrap functionality for cells"""

            def __init__(self, *args: Any, **kwargs: Any) -> None:
                super().__init__(*args, **kwargs)
                self.cell_height_ratio = 1.5  # Default cell height ratio

            def multi_cell_auto_width(
                self,
                width: float,
                height: float,
                text: str,
                border: int = 0,
                align: str = "L",
                fill: bool = False,
            ) -> float:
                """Create a multi-cell with automatic word wrapping and return its height"""
                # Calculate how many lines this text will require
                string_width: float = float(self.get_string_width(text))  # Explicitly cast to float
                lines_float: float = string_width / (width - 2)
                lines_count: int = max(1, int(lines_float) + 1)  # At least one line

                # Calculate total height needed
                total_height: float = height * float(lines_count)

                # Create the multi-cell
                self.multi_cell(width, height, text, border, align, fill)

                # Explicitly ensure we return a float
                return total_height

        # Store the class for later use
        PDFWithWordWrap_CLASS = PDFWithWordWrap

        PDF_CONVERTER = "fpdf"
    except ImportError:
        print("Warning: Neither win32com nor fpdf found. PDF conversion will be disabled.")
        PDF_CONVERTER = None


def create_output_folders(directory: str) -> None:
    """Create the output folders if they don't exist"""
    for folder in ["PDF", "CSV", "TXT"]:
        folder_path: str = os.path.join(directory, folder)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
            print(f"Created folder: {folder_path}")


def convert_to_csv(excel_file: str, output_folder: str) -> None:
    """Convert Excel file to CSV format"""
    filename: str = os.path.basename(excel_file)
    base_name: str = os.path.splitext(filename)[0]

    try:
        # Read all sheets from the Excel file using context manager
        with pd.ExcelFile(excel_file) as excel:
            sheet_names: List[str] = excel.sheet_names

            if len(sheet_names) == 1:
                # If there's only one sheet, convert it directly
                output_file: str = os.path.join(output_folder, f"{base_name}.csv")
                df_single: pd.DataFrame = pd.read_excel(excel, sheet_name=sheet_names[0])
                df_single.to_csv(output_file, index=False)
                print(f"Converted {excel_file} to CSV: {output_file}")
            else:
                # If there are multiple sheets, create separate CSV files
                for sheet in sheet_names:
                    sheet_output: str = os.path.join(output_folder, f"{base_name}_{sheet}.csv")
                    df_multi: pd.DataFrame = pd.read_excel(excel, sheet_name=sheet)
                    df_multi.to_csv(sheet_output, index=False)
                print(f"Converted {excel_file} to multiple CSV files in {output_folder}")

    except Exception as e:
        print(f"Error converting {excel_file} to CSV: {str(e)}")


def convert_to_txt(excel_file: str, output_folder: str) -> None:
    """Convert Excel file to TXT format"""
    filename: str = os.path.basename(excel_file)
    base_name: str = os.path.splitext(filename)[0]
    output_file: str = os.path.join(output_folder, f"{base_name}.txt")

    try:
        # Read all sheets from the Excel file using context manager
        with pd.ExcelFile(excel_file) as excel:
            sheet_names: List[str] = excel.sheet_names

            with open(output_file, "w", encoding="utf-8") as txt_file:
                if len(sheet_names) > 1:
                    # If there are multiple sheets, include sheet names in the TXT file
                    for sheet in sheet_names:
                        df_sheet: pd.DataFrame = pd.read_excel(excel, sheet_name=sheet)
                        txt_file.write(f"Sheet: {sheet}\n")
                        txt_file.write("=" * 50 + "\n")
                        txt_file.write(df_sheet.to_string(index=False))
                        txt_file.write("\n\n")
                else:
                    # If there's only one sheet, convert it directly
                    df_txt: pd.DataFrame = pd.read_excel(excel, sheet_name=sheet_names[0])
                    txt_file.write(df_txt.to_string(index=False))

        print(f"Converted {excel_file} to TXT: {output_file}")

    except Exception as e:
        print(f"Error converting {excel_file} to TXT: {str(e)}")


def convert_to_pdf_win32(excel_file: str, output_folder: str) -> None:
    """Convert Excel file to PDF format using win32com (Windows only)"""
    filename: str = os.path.basename(excel_file)
    base_name: str = os.path.splitext(filename)[0]
    output_file: str = os.path.join(output_folder, f"{base_name}.pdf")

    excel_app = None
    workbook = None
    try:
        # Initialize Excel application
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = False  # Run in background
        excel_app.DisplayAlerts = False  # Prevent Excel alerts

        # Open the workbook
        workbook = excel_app.Workbooks.Open(os.path.abspath(excel_file))

        # Process each worksheet
        for sheet in workbook.Worksheets:
            # Set to horizontal (landscape) orientation
            sheet.PageSetup.Orientation = 2  # xlLandscape

            # Fit to page width (1 page wide)
            sheet.PageSetup.FitToPagesWide = 1
            sheet.PageSetup.FitToPagesTall = False

            # Adjust column widths to fit content
            sheet.UsedRange.Columns.AutoFit()

            # Set print area to the used range
            sheet.PageSetup.PrintArea = sheet.UsedRange.Address

            # Set other print settings for better visibility
            sheet.PageSetup.Zoom = False  # Disable zoom to use FitToPages
            sheet.PageSetup.CenterHorizontally = True

        # Save as PDF
        workbook.ExportAsFixedFormat(0, os.path.abspath(output_file))

        print(f"Converted {excel_file} to PDF: {output_file}")

    except Exception as e:
        print(f"Error converting {excel_file} to PDF: {str(e)}")
    finally:
        # Always close workbook and Excel application to release file handles
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel_app is not None:
            try:
                excel_app.Quit()
            except Exception:
                pass


def convert_to_pdf_fpdf(excel_file: str, output_folder: str) -> None:
    """Convert Excel file to PDF format using fpdf with improved cell visibility"""
    # Check if PDFWithWordWrap_CLASS is available
    if PDFWithWordWrap_CLASS is None:
        print(f"Error: FPDF library not available for PDF conversion of {excel_file}")
        return

    filename: str = os.path.basename(excel_file)
    base_name: str = os.path.splitext(filename)[0]
    output_file: str = os.path.join(output_folder, f"{base_name}.pdf")

    wb = None
    try:
        # Load the Excel file
        wb = openpyxl.load_workbook(excel_file)

        # Create a PDF object - A4 landscape (horizontal)
        pdf = cast(Any, PDFWithWordWrap_CLASS)(orientation="L", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=15)

        # For each sheet in the workbook
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            pdf.add_page()

            # Add sheet name as a title
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, f"Sheet: {sheet_name}", 0, 1, "C")
            pdf.ln(5)

            # Get max rows and columns
            max_row: int = sheet.max_row
            max_col: int = sheet.max_column

            # Adjust for empty sheets
            if max_row == 0 or max_col == 0:
                pdf.set_font("Arial", "", 12)
                pdf.cell(0, 10, "Empty sheet", 0, 1, "C")
                continue

            # Calculate optimal column widths based on content
            col_widths: Dict[int, float] = {}
            max_content_lengths: Dict[int, int] = {}

            # First pass: determine maximum content length for each column
            for col in range(1, max_col + 1):
                max_length = 0
                for row in range(1, min(max_row + 1, 101)):  # Sample first 100 rows
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        max_length = max(max_length, len(str(cell_value)))
                max_content_lengths[col] = max_length

            # Available page width (A4 landscape minus margins)
            available_width: float = 277 - 20  # A4 landscape width ~277mm, minus margins

            # Calculate proportional column widths with minimum and maximum constraints
            total_content_length: int = sum(max_content_lengths.values())
            min_col_width: float = 10.0  # Minimum column width in mm
            max_col_width: float = 50.0  # Maximum column width in mm

            for col in range(1, max_col + 1):
                if total_content_length > 0:
                    proportion: float = max_content_lengths[col] / total_content_length
                    col_widths[col] = max(
                        min_col_width, min(max_col_width, proportion * available_width)
                    )
                else:
                    col_widths[col] = min_col_width

            # Adjust if total width exceeds available width
            total_width: float = sum(col_widths.values())
            if total_width > available_width:
                scale_factor: float = available_width / total_width
                for col in col_widths:
                    col_widths[col] *= scale_factor

            # Determine appropriate font size based on table size
            if max_col > 10 or max_row > 100:
                pdf.set_font("Arial", "B", 7)  # Smaller font for large tables
            elif max_col > 6 or max_row > 50:
                pdf.set_font("Arial", "B", 8)
            else:
                pdf.set_font("Arial", "B", 10)

            # Header row
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=1, column=col).value
                header_value_str: str = str(cell_value) if cell_value is not None else ""
                pdf.cell(col_widths[col], 8, header_value_str, 1, 0, "C")
            pdf.ln()

            # Data rows
            pdf.set_font("Arial", "", pdf.font_size - 1)  # Slightly smaller for data

            # Calculate how many rows to include
            max_rows_to_show: int = min(max_row, 1000)  # Limit to 1000 rows for large files

            for row in range(2, max_rows_to_show + 1):
                # Save current position
                start_x = pdf.get_x()
                start_y = pdf.get_y()
                max_height = 7  # Default row height

                # First pass: Calculate maximum height needed for this row
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    cell_text: str = str(cell_value) if cell_value is not None else ""

                    # Calculate cell height based on content
                    if len(cell_text) > 0:
                        cell_width = col_widths[col]
                        lines_needed = max(
                            1, int(pdf.get_string_width(cell_text) / (cell_width - 2)) + 1
                        )
                        height_needed = lines_needed * 5  # 5mm per line
                        max_height = max(max_height, height_needed)

                # Second pass: Draw cells with consistent height
                pdf.set_xy(start_x, start_y)
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    cell_content: str = str(cell_value) if cell_value is not None else ""

                    # For the last column in the row, use 1 for line break
                    line_break = 0 if col < max_col else 1

                    # Handle long text by using multi_cell if needed
                    if len(cell_content) > 15:  # Arbitrary threshold for "long" text
                        current_x = pdf.get_x()
                        current_y = pdf.get_y()

                        # Draw cell border first
                        pdf.cell(col_widths[col], max_height, "", 1, 0)

                        # Then draw text inside with proper positioning
                        pdf.set_xy(current_x, current_y)
                        pdf.multi_cell(col_widths[col], max_height / 2, cell_content, 0, "L")

                        # Reset position for next cell
                        pdf.set_xy(current_x + col_widths[col], current_y)
                    else:
                        pdf.cell(col_widths[col], max_height, cell_content, 1, line_break, "L")

                # Check if we need a page break
                if pdf.get_y() > pdf.h - 20:  # 20mm from bottom of page
                    pdf.add_page()

            # If there are more rows than we can show
            if max_row > max_rows_to_show:
                pdf.ln(5)
                pdf.set_font("Arial", "I", 8)
                pdf.cell(
                    0,
                    10,
                    f"Note: {max_row - max_rows_to_show} additional rows not shown",
                    0,
                    1,
                    "C",
                )

        # Save the PDF
        pdf.output(output_file)
        print(f"Converted {excel_file} to PDF: {output_file}")

    except Exception as e:
        print(f"Error converting {excel_file} to PDF: {str(e)}")
    finally:
        # Always close the workbook to release file handles
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def convert_excel_file(excel_file: str) -> None:
    """Convert an Excel file to multiple formats"""
    directory: str = os.path.dirname(excel_file)

    # Create output folders
    create_output_folders(directory)

    # Convert to different formats
    convert_to_csv(excel_file, os.path.join(directory, "CSV"))
    convert_to_txt(excel_file, os.path.join(directory, "TXT"))

    # Convert to PDF if a converter is available
    if PDF_CONVERTER == "win32":
        convert_to_pdf_win32(excel_file, os.path.join(directory, "PDF"))
    elif PDF_CONVERTER == "fpdf":
        convert_to_pdf_fpdf(excel_file, os.path.join(directory, "PDF"))


def process_directory(directory: str) -> None:
    """Process a directory, looking for Excel files"""
    excel_found: bool = False

    # Walk through the directory and its subdirectories
    for root, dirs, files in os.walk(directory):
        excel_files: List[str] = [
            os.path.join(root, file) for file in files if file.lower().endswith(".xlsx")
        ]

        if excel_files:
            excel_found = True
            print(f"Found {len(excel_files)} Excel files in {root}")

            for excel_file in excel_files:
                convert_excel_file(excel_file)

    if not excel_found:
        print(f"No Excel (.xlsx) files found in {directory} or its subdirectories.")


def main() -> None:
    """Main function to handle command line arguments"""
    parser = argparse.ArgumentParser(
        description="Convert Excel files to PDF, CSV, and TXT formats."
    )
    parser.add_argument(
        "directory",
        nargs="?",
        default=os.getcwd(),
        help="Directory to scan for Excel files (default: current directory)",
    )
    args = parser.parse_args()

    # Validate directory
    if not os.path.isdir(args.directory):
        print(f"Error: {args.directory} is not a valid directory")
        sys.exit(1)

    print(f"Processing directory: {args.directory}")
    process_directory(args.directory)
    print("Processing complete.")


if __name__ == "__main__":
    main()
